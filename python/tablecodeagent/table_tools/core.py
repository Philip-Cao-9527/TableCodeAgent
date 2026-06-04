from __future__ import annotations

import csv
import datetime as dt
import statistics
from pathlib import Path
from typing import Any, Iterable


def _stringify_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    return str(value)


def _make_unique_columns(columns: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    unique: list[str] = []
    for idx, column in enumerate(columns, start=1):
        base = column.strip() or f"column_{idx}"
        count = seen.get(base, 0) + 1
        seen[base] = count
        unique.append(base if count == 1 else f"{base}_{count}")
    return unique


def _normalize_header_rows(header_grid: list[list[str]]) -> list[str]:
    if not header_grid:
        return []

    width = max(len(row) for row in header_grid)
    filled_rows: list[list[str]] = []
    for row in header_grid:
        padded = row + [""] * (width - len(row))
        filled: list[str] = []
        last = ""
        for value in padded:
            current = value.strip()
            if current:
                last = current
            filled.append(current or last)
        filled_rows.append(filled)

    columns: list[str] = []
    for col_idx in range(width):
        parts: list[str] = []
        for row in filled_rows:
            part = row[col_idx].strip()
            if part and (not parts or parts[-1] != part):
                parts.append(part)
        columns.append("__".join(parts))
    return _make_unique_columns(columns)


def _rows_from_grid(grid: list[list[str]], header_rows: int = 1) -> tuple[list[str], list[dict[str, str]]]:
    if header_rows < 1:
        raise ValueError("header_rows must be >= 1.")
    if len(grid) < header_rows:
        return [], []

    width = max((len(row) for row in grid), default=0)
    normalized_grid = [row + [""] * (width - len(row)) for row in grid]
    columns = _normalize_header_rows(normalized_grid[:header_rows])
    rows: list[dict[str, str]] = []
    for values in normalized_grid[header_rows:]:
        if not any(_stringify_cell(value).strip() for value in values):
            continue
        rows.append({columns[idx]: _stringify_cell(values[idx]) for idx in range(len(columns))})
    return columns, rows


def _read_csv_grid(path: Path) -> list[list[str]]:
    with path.open("r", encoding="utf-8", newline="") as f:
        return [[_stringify_cell(cell) for cell in row] for row in csv.reader(f)]


def _read_xlsx_grid(
    path: Path,
    sheet_name: str | None = None,
    fill_merged_cells: bool = False,
) -> tuple[list[list[str]], str, list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise ImportError("Reading .xlsx files requires openpyxl.") from error

    workbook = load_workbook(path, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    grid = [[_stringify_cell(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]

    if fill_merged_cells:
        for merged_range in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            value = grid[min_row - 1][min_col - 1] if min_row - 1 < len(grid) else ""
            for row_idx in range(min_row - 1, max_row):
                while row_idx >= len(grid):
                    grid.append([])
                while max_col > len(grid[row_idx]):
                    grid[row_idx].append("")
                for col_idx in range(min_col - 1, max_col):
                    grid[row_idx][col_idx] = value

    return grid, sheet.title, list(workbook.sheetnames)


def _read_table(
    table_path: str | Path,
    *,
    sheet_name: str | None = None,
    header_rows: int = 1,
    fill_merged_cells: bool = False,
) -> tuple[list[str], list[dict[str, str]], dict[str, Any]]:
    path = Path(table_path)
    suffix = path.suffix.lower()
    metadata: dict[str, Any] = {"path": str(path), "format": suffix.lstrip(".") or "csv"}

    if suffix in ("", ".csv"):
        grid = _read_csv_grid(path)
    elif suffix == ".xlsx":
        grid, actual_sheet, sheet_names = _read_xlsx_grid(path, sheet_name, fill_merged_cells)
        metadata["sheet_name"] = actual_sheet
        metadata["sheet_names"] = sheet_names
        metadata["fill_merged_cells"] = fill_merged_cells
    else:
        raise ValueError(f"Unsupported table format: {suffix}")

    metadata["header_rows"] = header_rows
    columns, rows = _rows_from_grid(grid, header_rows)
    return columns, rows, metadata


def _read_rows(
    csv_path: str | Path,
    *,
    sheet_name: str | None = None,
    header_rows: int = 1,
    fill_merged_cells: bool = False,
) -> list[dict[str, str]]:
    return _read_table(
        csv_path,
        sheet_name=sheet_name,
        header_rows=header_rows,
        fill_merged_cells=fill_merged_cells,
    )[1]


def _to_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _is_missing(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _read_header(
    csv_path: str | Path,
    *,
    sheet_name: str | None = None,
    header_rows: int = 1,
    fill_merged_cells: bool = False,
) -> list[str]:
    return _read_table(
        csv_path,
        sheet_name=sheet_name,
        header_rows=header_rows,
        fill_merged_cells=fill_merged_cells,
    )[0]


def _infer_type(values: Iterable[str]) -> str:
    present = [v for v in values if not _is_missing(v)]
    if not present:
        return "empty"

    nums = [_to_number(v) for v in present]
    if all(v is not None for v in nums):
        return "integer" if all(float(v).is_integer() for v in nums if v is not None) else "number"

    try:
        for value in present:
            dt.date.fromisoformat(str(value))
        return "date"
    except ValueError:
        return "string"


def _column_numeric_stats(values: list[str]) -> dict[str, float | int] | None:
    nums = [_to_number(value) for value in values]
    nums = [value for value in nums if value is not None]
    if not nums:
        return None
    return {
        "count": len(nums),
        "min": min(nums),
        "max": max(nums),
        "mean": statistics.fmean(nums),
        "sum": sum(nums),
    }


def load_table(
    csv_path: str | Path,
    preview_rows: int = 5,
    sheet_name: str | None = None,
    header_rows: int = 1,
    fill_merged_cells: bool = False,
) -> dict[str, Any]:
    columns, rows, metadata = _read_table(
        csv_path,
        sheet_name=sheet_name,
        header_rows=header_rows,
        fill_merged_cells=fill_merged_cells,
    )
    preview_size = max(0, preview_rows)
    return {
        **metadata,
        "columns": columns,
        "row_count": len(rows),
        "preview_row_count": min(preview_size, len(rows)),
        "preview_rows": rows[:preview_size],
    }


def profile_table(
    csv_path: str | Path,
    sheet_name: str | None = None,
    header_rows: int = 1,
    fill_merged_cells: bool = False,
) -> dict[str, Any]:
    columns, rows, metadata = _read_table(
        csv_path,
        sheet_name=sheet_name,
        header_rows=header_rows,
        fill_merged_cells=fill_merged_cells,
    )

    missing = {col: sum(1 for row in rows if row.get(col, "") == "") for col in columns}
    numeric_stats: dict[str, dict[str, float | int]] = {}
    column_profiles: dict[str, dict[str, Any]] = {}

    for col in columns:
        values = [row.get(col, "") for row in rows]
        stats = _column_numeric_stats(values)
        if stats:
            numeric_stats[col] = stats

        missing_count = missing[col]
        col_flags: list[str] = []
        if rows and missing_count:
            col_flags.append("has_missing")
        if rows and missing_count == len(rows):
            col_flags.append("all_missing")

        column_profiles[col] = {
            "missing_count": missing_count,
            "missing_rate": (missing_count / len(rows)) if rows else 0.0,
            "unique_count": len({value for value in values if not _is_missing(value)}),
            "inferred_type": _infer_type(values),
            "numeric_stats": stats,
            "flags": col_flags,
        }

    row_signatures = [tuple(row.get(col, "") for col in columns) for row in rows]
    duplicate_row_count = len(row_signatures) - len(set(row_signatures))
    quality_flags: list[str] = []
    if not rows:
        quality_flags.append("empty_table")
    if duplicate_row_count:
        quality_flags.append("duplicate_rows")
    if any(count > 0 for count in missing.values()):
        quality_flags.append("columns_with_missing")

    return {
        **metadata,
        "row_count": len(rows),
        "column_count": len(columns),
        "columns": columns,
        "missing_values": missing,
        "numeric_stats": numeric_stats,
        "column_profiles": column_profiles,
        "duplicate_row_count": duplicate_row_count,
        "quality_flags": quality_flags,
    }


def _normalize_filters(filters: Any) -> list[dict[str, Any]]:
    if not filters:
        return []
    if isinstance(filters, dict):
        if {"column", "op", "value"}.issubset(filters):
            return [filters]
        return [{"column": key, "op": "eq", "value": value} for key, value in filters.items()]
    if isinstance(filters, list):
        normalized = []
        for item in filters:
            if not isinstance(item, dict):
                raise ValueError("Each filter must be an object.")
            if "column" not in item or "value" not in item:
                raise ValueError("Each filter must include column and value.")
            normalized.append({
                "column": item["column"],
                "op": item.get("op", "eq"),
                "value": item["value"],
            })
        return normalized
    raise ValueError("filters must be an object or a list of objects.")


def _matches_filter(row: dict[str, str], condition: dict[str, Any]) -> bool:
    column = condition["column"]
    op = condition.get("op", "eq")
    expected = condition["value"]
    actual = row.get(column, "")

    if op in ("eq", "=="):
        return str(actual) == str(expected)
    if op in ("ne", "!="):
        return str(actual) != str(expected)
    if op == "contains":
        return str(expected) in str(actual)

    actual_num = _to_number(actual)
    expected_num = _to_number(expected)
    if actual_num is None or expected_num is None:
        raise ValueError(f"Filter {op} requires numeric values for column: {column}")

    if op in ("gt", ">"):
        return actual_num > expected_num
    if op in ("gte", ">="):
        return actual_num >= expected_num
    if op in ("lt", "<"):
        return actual_num < expected_num
    if op in ("lte", "<="):
        return actual_num <= expected_num

    raise ValueError(f"Unsupported filter op: {op}")


def query_table(
    csv_path: str | Path,
    metric: str,
    column: str | None = None,
    filters: Any = None,
    sheet_name: str | None = None,
    header_rows: int = 1,
    fill_merged_cells: bool = False,
) -> dict[str, Any]:
    rows = _read_rows(
        csv_path,
        sheet_name=sheet_name,
        header_rows=header_rows,
        fill_merged_cells=fill_merged_cells,
    )
    normalized_filters = _normalize_filters(filters)

    matched = [
        row for row in rows
        if all(_matches_filter(row, condition) for condition in normalized_filters)
    ]

    if metric == "count":
        value: Any = len(matched)
    else:
        if not column:
            raise ValueError(f"metric {metric} requires column.")
        nums = [_to_number(row.get(column, "")) for row in matched]
        nums = [v for v in nums if v is not None]

        if metric == "sum":
            value = sum(nums)
        elif metric == "mean":
            value = statistics.fmean(nums) if nums else None
        elif metric == "max":
            value = max(nums) if nums else None
        elif metric == "min":
            value = min(nums) if nums else None
        else:
            raise ValueError(f"Unsupported metric: {metric}")

    return {
        "value": value,
        "metric": metric,
        "column": column,
        "matched_row_count": len(matched),
        "total_row_count": len(rows),
        "filters": normalized_filters,
        "basis": {
            "operation": metric,
            "target_column": column,
            "filter_count": len(normalized_filters),
        },
    }


def _table_spec_path(spec: dict[str, Any]) -> str:
    return spec.get("path") or spec.get("csv_path") or spec.get("table_path")


def _read_named_table(spec: dict[str, Any]) -> tuple[str, list[dict[str, str]]]:
    name = spec["name"]
    path = _table_spec_path(spec)
    if not path:
        raise ValueError(f"Table spec {name} must include path.")
    rows = _read_rows(
        path,
        sheet_name=spec.get("sheet_name"),
        header_rows=int(spec.get("header_rows", 1)),
        fill_merged_cells=bool(spec.get("fill_merged_cells", False)),
    )
    return name, rows


def query_multi_table(
    tables: list[dict[str, Any]],
    join: dict[str, Any],
    metric: str,
    column: str | None = None,
    filters: Any = None,
) -> dict[str, Any]:
    if len(tables) != 2:
        raise ValueError("query_multi_table currently supports exactly two tables.")

    left_name, left_rows = _read_named_table(tables[0])
    right_name, right_rows = _read_named_table(tables[1])
    left_key = join["left_key"]
    right_key = join["right_key"]

    right_index: dict[str, list[dict[str, str]]] = {}
    for row in right_rows:
        right_index.setdefault(row.get(right_key, ""), []).append(row)

    joined_rows: list[dict[str, str]] = []
    for left_row in left_rows:
        for right_row in right_index.get(left_row.get(left_key, ""), []):
            combined: dict[str, str] = {}
            for key, value in left_row.items():
                combined[f"{left_name}.{key}"] = value
            for key, value in right_row.items():
                combined[f"{right_name}.{key}"] = value
            joined_rows.append(combined)

    normalized_filters = _normalize_filters(filters)
    matched = [
        row for row in joined_rows
        if all(_matches_filter(row, condition) for condition in normalized_filters)
    ]

    if metric == "count":
        value: Any = len(matched)
    else:
        if not column:
            raise ValueError(f"metric {metric} requires column.")
        nums = [_to_number(row.get(column, "")) for row in matched]
        nums = [v for v in nums if v is not None]

        if metric == "sum":
            value = sum(nums)
        elif metric == "mean":
            value = statistics.fmean(nums) if nums else None
        elif metric == "max":
            value = max(nums) if nums else None
        elif metric == "min":
            value = min(nums) if nums else None
        else:
            raise ValueError(f"Unsupported metric: {metric}")

    return {
        "value": value,
        "metric": metric,
        "column": column,
        "matched_row_count": len(matched),
        "joined_row_count": len(joined_rows),
        "filters": normalized_filters,
        "basis": {
            "operation": metric,
            "target_column": column,
            "filter_count": len(normalized_filters),
            "join": {
                "left_table": left_name,
                "right_table": right_name,
                "left_key": left_key,
                "right_key": right_key,
            },
        },
    }
